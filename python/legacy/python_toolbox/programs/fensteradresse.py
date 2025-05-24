from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from PyPDF2 import PdfReader, PdfWriter
import PyPDF2
import os
from .utils import printlb, count_pdfs, get_user_input


def excel_zu_bestehender_pdf(excel_datei, pdf_datei, komm_id, output_datei, return_adress):

    # load xlsx-table
    wb = openpyxl.load_workbook(excel_datei)
    ws = wb.active

    # load_pdf
    existing_pdf = PdfReader(open(pdf_datei, "rb"))
    page = existing_pdf.pages[0] 

    # create temporary pdf
    temp_pdf_datei = "temp_with_address.pdf"
    c = canvas.Canvas(temp_pdf_datei, pagesize=page.mediabox.upper_right)

    # fensteradresse position
    x1              = 40  
    y_return_adress = 690
    y_line          = 685
    y_name          = 660  
    y_street        = 645  
    y_postal        = 630  
    y_salutation    = 573 

    # Flag to track if PDF was found by name
    found = False

    # Daten aus Excel lesen und in PDF schreiben, basierend auf komm_id
    for row in ws.iter_rows(min_row=2, values_only=True): 
        if row[0] == komm_id:
            _, anrede, name, vorname, strasse, hausnummer, plz, ort = row

            if return_adress is None:
                return_adress = "IMIBE UK-Essen, Hufelandstr. 55, 45147 Essen"

            c.setFont("Helvetica-Oblique", 9)
            c.drawString(x1, y_return_adress, return_adress)

            begrüßung = ""
            if anrede == "M":
                anrede = "Herr"
                begrüßung = "Sehr geehrter Herr"
            elif anrede == "W":
                begrüßung = "Sehr geehrte Frau"
                anrede = "Frau"
            else:
                anrede = ""
                begrüßung = "Sehr geehrte/r Herr/Frau"

            # Zurücksetzen auf normale Schriftart für die restlichen Zeilen
            c.setFont("Helvetica", 11)
            trenner = "_________________________________"
            c.drawString(x1, y_line, trenner)
            adresse1 = f"{anrede} {vorname} {name}"
            c.drawString(x1, y_name, adresse1)
            adresse2 = f"{strasse} {hausnummer}"
            c.drawString(x1, y_street, adresse2)
            adresse3 = f"{plz} {ort}"
            c.drawString(x1, y_postal, adresse3)
            begrüßung = f"{begrüßung} {name},"
            c.drawString(x1, y_salutation, begrüßung)

            found = True
            break

    # Print whether komm_id was found
    if found:
        pass
    else:
        print(f"komm_id {komm_id} nicht gefunden.")

    # Seite zur neuen PDF hinzufügen
    c.showPage()
    c.save()

    try:
        # Bestehende und neue PDF zusammenführen
        output = PdfWriter()

        # Laden der temporären PDF mit der Adresse
        temp_pdf = PdfReader(open(temp_pdf_datei, "rb"))
        # Hinzufügen der ersten Seite (mit der Adresse) zur Ausgabe
        output.add_page(temp_pdf.pages[0])

        # Hinzufügen der restlichen Seiten der bestehenden PDF
        for i in range(len(existing_pdf.pages)):
            page = existing_pdf.pages[i]
            output.add_page(page)

        # Ausgabe speichern
        with open(output_datei, "wb") as outputStream:
            output.write(outputStream)

        # Print success message
        print(f"Fensteradresse erfolgreich hinzugefügt! -- {name}, {vorname}")

    except Exception as e:
        print(f"Fehler beim Hinzufügen der Fensteradresse: {e}")


def overlay_first_two_pages(input_pdf, output_pdf):
    with open(input_pdf, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        pdf_writer = PyPDF2.PdfWriter()

        if len(pdf_reader.pages) < 2:
            raise ValueError("Die PDF-Datei muss mindestens zwei Seiten enthalten.")

        page1 = pdf_reader.pages[0]
        page2 = pdf_reader.pages[1]

        # Die erste Seite über die zweite legen
        page2.merge_page(page1)

        # Die modifizierte zweite Seite hinzufügen
        pdf_writer.add_page(page2)

        # Alle restlichen Seiten hinzufügen
        for page_num in range(2, len(pdf_reader.pages)):
            pdf_writer.add_page(pdf_reader.pages[page_num])

        with open(output_pdf, 'wb') as output_file:
            pdf_writer.write(output_file)


def delete_pdf_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
    else:
        print(f"Die Datei konnte nicht gelöscht werden, weil '{file_path}' existiert nicht.")


def extract_first_column_as_strings(dateipfad):
    workbook = openpyxl.load_workbook(dateipfad)
    worksheet = workbook.active

    first_column = []
    for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True): 
        value = row[0]
        if value is not None:
            first_column.append('F' + str(value)) 

    return first_column


def add_fensteradresse_to_all(input_dir, xlsx_table, output_dir, return_adress):
    pdfs_not_found = []
    pdf_names_in_table = extract_first_column_as_strings(xlsx_table)
    pdfs_in_input_dir = count_pdfs(input_dir)
    pdfs_moved_to_output_dir = 0

    for x in range(len(pdf_names_in_table)):
        temp_pdf = f"{input_dir}\\{pdf_names_in_table[x]}.pdf"
        temp_pdf_zs = f"{input_dir}\\{pdf_names_in_table[x]}_zs.pdf"
        temp_pdf_output = f"{output_dir}\\{pdf_names_in_table[x]}_af.pdf"

        if os.path.exists(temp_pdf):
            print(f"\nPDF '{pdf_names_in_table[x]}' gefunden!")
            excel_zu_bestehender_pdf(xlsx_table, temp_pdf, int(pdf_names_in_table[x][1:]), temp_pdf_zs, return_adress)
            overlay_first_two_pages(temp_pdf_zs, temp_pdf_output)
            pdfs_moved_to_output_dir += 1

        else:
            pdfs_not_found.append(temp_pdf)

    if pdfs_not_found != []:
        print("\nDie folgenden PDFs konnten nicht gefunden werden:")
        for x in range (len(pdfs_not_found)):
            print(f"[{x+1}] {pdfs_not_found[x]}")
    else:
        print("\nAlle PDFs in der XLSX-Tabelle konnten in dem Ordner gefunden werden.")

    print(f"\n{pdfs_moved_to_output_dir} PDFs wurden verschoben nach: {output_dir}")
    print(f"\n{pdfs_in_input_dir-pdfs_moved_to_output_dir} PDFs verbleiben in: {input_dir}")



def run_this_program():
    input_dir  = get_user_input("path", "Gib den Pfad des Ordners ein, der die zu bearbeitenden PDFs enthält")
    xlsx_table = get_user_input("path", "Gib den Pfad der XLSX-Tabelle ein, welche die Adressen enthält")
    output_dir = get_user_input("path", "Gib den Pfad des Ordners ein, in dem die bearbeiteten PDFs abgelegt werden sollen")
    while True:
        change_return_adress = get_user_input("bool", "Die Rücksendeadresse lautet standardmäßig:\n'IMIBE UK-Essen, Hufelandstr. 55, 45147 Essen'\nMöchtest du das ändern?")
        if change_return_adress:
            return_adress = get_user_input("string", "Gib die Rücksendeadresse ein, die du verwenden willst")
            if len(return_adress) >= 45:
                approve_return_adress = get_user_input("bool", "Diese Rücksendeadresse ist etwas länger als gewöhnlich. Dies könnte bedeuten, dass sie im Adressfenster nicht richtig zu sehen ist.\nWillst du trotzdem fortfahren?")
                if approve_return_adress:
                    break
                
        else:
            return_adress = None
            break

    add_fensteradresse_to_all(input_dir, xlsx_table, output_dir, return_adress)

    input("Drücke Enter, um fortzufahren...")


help_text = ["Das Programm automatisiert das Hinzufügen von Adressinformationen zu PDF-Dokumenten. Es liest Adressen aus einer",
             "XLSX-Tabelle aus und fügt sie als Fensteradresse die PDFs ein. Die Adressinformationen werden so positioniert, dass sie",
             "in einem Adressfenster gut zu sehen sind. Dafür musst du nur 3 Informationen angeben:",
             "(1) Den Dateipfad an dem sich die zu bearbeitenden PDFs befinden",
             "(2) Die XLSX-Tabelle in der sich die Adressinformationen befinden und",
             "(3) Den Dateipfad in dem die bearbeiteten Dateien abgelegt werden sollen",
             "\nHinweis: Die Tabelle muss folgendermaßen formatiert sein:",
             "dateiname | anrede |  name   | vorname | strasse   | hausnummer | plz   | ort",
             "F12345    | Herr   | Caesar  | Julius  | Romstraße | 12345      | 00184 | roma",]
    